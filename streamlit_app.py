import streamlit as st
import pandas as pd
import openpyxl
from itertools import combinations
from openpyxl.styles import PatternFill
from io import BytesIO

TOL = 3  # tolerancia ±

st.set_page_config("Pareo Óptimo", layout="wide")
st.title("🔍 Resaltado de Pares en Excel")

@st.cache_data
def load_df(uploaded):
    return pd.read_excel(uploaded)

def clean(df):
    first = df.columns[0]
    df = df[~df[first].astype(str).str.endswith("Total")]
    df = df.drop(columns=[first]).reset_index(drop=True)
    return df

def find_pairs(df, col4, col8):
    rows = []
    for _, grp in df.groupby(col4):
        idxs = list(grp.index)
        vals = grp[col8]
        for i, j in combinations(idxs, 2):
            s = vals[i] + vals[j]
            if abs(s) <= TOL:
                rows.append({"i": i, "j": j, "suma": s})
    return pd.DataFrame(rows)

def greedy_pairs(pairs_df):
    sel, used = set(), set()
    if pairs_df.empty:
        return sel
    df = pairs_df.assign(abs_sum=lambda x: x["suma"].abs())
    for _, r in df.sort_values("abs_sum").iterrows():
        i, j = int(r.i), int(r.j)
        if i not in used and j not in used:
            sel |= {i, j}
            used |= {i, j}
    return sel

def to_excel_highlight(df, idxs):
    buf = BytesIO()
    df.to_excel(buf, index=False, sheet_name="Datos")
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    yellow = PatternFill("solid", fgColor="FFFF00")
    for row in ws.iter_rows(min_row=2):
        idx = row[0].row - 2
        if idx in idxs:
            for cell in row:
                cell.fill = yellow
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

uploaded = st.file_uploader("1️⃣ Sube tu Excel (.xlsx)", type="xlsx")
if uploaded:
    df = load_df(uploaded)
    st.subheader("Datos Originales")
    st.dataframe(df, use_container_width=True)

    df_clean = clean(df)
    st.subheader("Datos Limpios (sin 'Total' y sin 1ª col.)")
    st.dataframe(df_clean, use_container_width=True)

    col4, col8 = df.columns[3], df.columns[7]
    pairs = find_pairs(df_clean, col4, col8)
    st.subheader(f"Pares con suma en ±{TOL}")
    st.dataframe(pairs, use_container_width=True)

    idxs = greedy_pairs(pairs)
    st.subheader(f"Filas resaltadas: {len(idxs)}")
    result_xlsx = to_excel_highlight(df_clean, idxs)

    st.download_button(
        "📥 Descargar Excel resaltado",
        result_xlsx,
        "pareos_resaltados.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
